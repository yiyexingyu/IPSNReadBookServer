# -*- encoding: utf-8 -*-
# @File    : allBookInfo.py
# @Time    : 2020/3/13 23:19
# @Author  : 一叶星羽
# @Email   : h0670131005@gmail.com
# @Software: PyCharm

import time
import re

from lxml import etree
import requests
from urllib import request


def getNovelInfo(novelUrl: str, filterRe: str, process=None) -> list:
    """
    根据小说的url地址和匹配规则filterRe，爬取小说的相关信息。process是额外处理函数，
    爬取相关信息后，对这些信息进行额外的处理。比如补全小说的每个章节的url。
    :param novelUrl: 小说相关信息的url
    :param filterRe: 相关信息匹配规则-制作表达式
    :param process: 额外的处理函数,必须接收一个参数。
    """

    # 向服务器发送请求，并获取返回的html页面
    html = request.urlopen(novelUrl).read().decode("utf-8")
    # 匹配的正则表达式
    novelInfo = re.findall(filterRe, html)

    if process:
        novelInfo = process(novelInfo)
    return novelInfo


def novelFromBiquge():
    url1 = "http://www.xbiquge.la/15/15003/"
    req = "<dd><a href='(.*?)' >(.*?)</a></dd>"

    chapters = getNovelInfo(url1, req)
    return chapters


def getNovelListByOriginName(originName: str):
    originInfo = {
        "笔趣阁":"http://www.xbiquge.la/xiaoshuodaquan/",
    }


def getNovelListFromBiquge(novelType=-1, start = 0, novelCount: int = 0, process=None) -> list:
    url = "http://www.xbiquge.la/xiaoshuodaquan/"

    # 向服务器发送请求，并获取返回的html页面
    # st = time.time()
    html = requests.get(url)
    # print("请求全部小说页面：", time.time() - st, "s")

    novelInfoList = []

    novelHtmlList = etree.HTML(html.text)

    novelTags = novelHtmlList.xpath("//*[@id='main']/div[@class='novellist']/h2/text()")
    novelTyped = novelHtmlList.xpath("//*[@id='main']/div[@class='novellist']")
    novelId = 0

    novels = {}
    if 0 <= novelType < len(novelTags):
        # novelTag = re.findall("(.*?)小说", novelTags[novelType].replace('、', ''))
        novelLinks = novelTyped[novelType].xpath("./ul/li/a/@href")
        novels[novelType] = novelLinks
    else:
        for i in range(len(novelTags)):
            # novelTag = re.findall("(.*?)小说", novelTags[i].replace('、', ''))
            novelLinks = novelTyped[i].xpath("./ul/li/a/@href")
            novels[i] = novelLinks


    for i in novels:
        novelTag = re.findall("(.*?)小说", novelTags[i].replace('、', ''))
        novelLinks = novels[i]

        # 爬取每一本小说的详细信息
        for n in range(start, len(novelLinks)):
            # st = time.time()
            novelLink = novelLinks[n]
            infoHtml = requests.get(novelLink)
            # print("请求小说详细页面：", time.time() - st, "s")

            infoHtml.encoding = infoHtml.apparent_encoding
            novelInfoHtml = etree.HTML(infoHtml.text)

            novelInfo = {"tag": ";".join(novelTag), "link": novelLink}
            novelInfo["title"] = novelInfoHtml.xpath("//*[@id='info']/h1/text()")[0]
            novelInfo["author"] = novelInfoHtml.xpath("//*[@id='info']/p[1]/text()")[0].split("：")[-1]
            novelInfo["state"] = "完结" # 需要额外处理

            intro = novelInfoHtml.xpath("//*[@id='intro']/p[2]/text()")
            if intro and len(intro) > 0:
                intro = intro[0].strip()
            else:
                intro = "无简介，这位作者很懒！！"
            novelInfo["intro"] = intro

            novelInfo["update"] = novelInfoHtml.xpath("//*[@id='info']/p[3]/text()")[0].split("：")[-1]
            novelInfo["cover"] = novelInfoHtml.xpath("//*[@id='fmimg']/img/@src")[0]

            novelInfoList.append(novelInfo)
            novelId += 1

            if novelCount > 0 and novelId >= novelCount:
                if process is not None:
                    process(novelInfoList)
                return novelInfoList

    if process is not None:
        process(novelInfoList)
    return novelInfoList


if __name__ == '__main__':
    from openpyxl.workbook import Workbook

    def saveDateToExcel(data: [dict]):
        wb = Workbook()

        ws = wb.worksheets[0]
        ws.title = "novel info"

        # 设置表头
        headRow = list(data[0].keys())
        for i in range(1, len(headRow) + 1):
            ws.cell(1, i, value=headRow[i-1])

        # 设置数据
        for index in range(0, len(data)):
            for col in range(0, len(headRow)):
                ws.cell(row=index+2, column=col+1, value=data[index][headRow[col]])

        wb.save("navelInfo.xls")


    novelInfo = []
    try:
        for i in range(0, 6):
                novelInfo.extend(getNovelListFromBiquge(novelType=i, novelCount=20, start=0))
    except Exception as e:
        print(e)
    finally:
        saveDateToExcel(novelInfo)
